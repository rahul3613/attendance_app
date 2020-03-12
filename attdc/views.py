from django.shortcuts import render,get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from .models import Instructor, Batch, Student, Atdc, S_code, PassWord
from django.utils import timezone
from uuid import getnode as get_mac
from getmac import get_mac_address
import openpyxl


def new(request):
    return render(request, 'attdc/new.html',)

def index(request):
    try:
        student = Student.objects.get(mac = user_mac(request))
    except (KeyError, Student.DoesNotExist):
        return render(request, 'attdc/index.html',)
    else:
        return HttpResponseRedirect(reverse('attdc:con_attdc', args=(student.id,)))


def get_batch1(request):
    instructor = get_object_or_404(Instructor, id=1)
    return render(request, 'attdc/get_batch1.html', {'instructor': instructor})

def get(request):
    try:
        check = PassWord.objects.get(passw = request.POST['pwd'])
    except (KeyError, PassWord.DoesNotExist):
        return render(request, 'attdc/unath.html',)
    else:
        instructor = get_object_or_404(Instructor, id=1)
        batch = instructor.batch_set.get(id=request.POST['batch'])
        return render(request, 'attdc/get.html', {'batch': batch})



def get_batch2(request):
    instructor = get_object_or_404(Instructor, id=1)
    return render(request, 'attdc/get_batch2.html', {'instructor': instructor})

def today(request):
    try:
        check = PassWord.objects.get(passw = request.POST['pwd'])
    except (KeyError, PassWord.DoesNotExist):
        return render(request, 'attdc/unath.html',)
    else:
        instructor = get_object_or_404(Instructor, id=1)
        batch = instructor.batch_set.get(id=request.POST['batch'])
        return render(request, 'attdc/get_tod.html', {'batch': batch})
    


def con_attdc(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'attdc/con_attdc.html', {'student': student})

def success(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'attdc/success.html', {'student': student})

def fail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'attdc/fail.html', {'student': student})

'''def ip_fail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'attdc/ip_fail.html', {'student': student})
'''


def mark(request):
    try:
        student = get_object_or_404(Student, rol_numb = request.POST['r_numb'])
    except (KeyError, Student.DoesNotExist):
        return render(request, 'attdc/index.html',)
    else:        
        if student.accept():
            student.tot_attdc += 1
            student.tod_date = timezone.now()
            student.mac = user_mac(request)
            student.save()
            return HttpResponseRedirect(reverse('attdc:success', args=(student.id,)))
        else:
            return HttpResponseRedirect(reverse('attdc:fail', args=(student.id,)))



def mark1(request):
    try:
        student = Student.objects.get(mac = user_mac(request))
    except (KeyError, Student.DoesNotExist):
        return render(request, 'attdc/index.html',)
    else:    
        if student.accept():
            student.tot_attdc += 1
            student.tod_date = timezone.now()
            student.save()
            return HttpResponseRedirect(reverse('attdc:success', args=(student.id,)))
        else:
            return HttpResponseRedirect(reverse('attdc:fail', args=(student.id,)))



def get_batch(request):
    instructor = get_object_or_404(Instructor, id=1)
    return render(request, 'attdc/get_batch.html', {'instructor': instructor})

def to_up(request):
    try:
        check = PassWord.objects.get(passw = request.POST['pwd'])
    except (KeyError, PassWord.DoesNotExist):
        return render(request, 'attdc/unath.html',)
    else:
        if S_code.objects.all():
            pass
        else:
            batch1 = S_code(code = 'M101')
            batch1.save()

        instructor = get_object_or_404(Instructor, id=1)
        batch2 = instructor.batch_set.get(id=request.POST['batch'])
        #batch = get_object_or_404(Batch, pk = request.POST.get('batch'))
    
        temp_code = get_object_or_404(S_code, id = 1)
        temp_code.code = batch2.s_code
        temp_code.save()

        return HttpResponseRedirect(reverse('attdc:up_file',))



def up_file(request):
    
    if request.method == 'POST':
        uploaded_file = request.FILES['myfile']
        print(uploaded_file.name)
        print(uploaded_file.size)

        wb_obj = openpyxl.load_workbook(uploaded_file)
        
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        batch1 = get_object_or_404(S_code, id = 1)
        batch = get_object_or_404(Batch, s_code = batch1.code)
    
        for i in range(2, m_row + 1): 
            name = sheet_obj.cell(row = i, column = 2)
            r_numb = sheet_obj.cell(row = i, column = 3)
        
            student = Student(batch= batch, rol_numb=int(r_numb.value), name=str(name.value))
            student.save()
        return render(request, 'attdc/up_succ.html',{'batch': batch})

    return render(request, 'attdc/up_file.html',)







def user_ip(request):

    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')

    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    #return HttpResponse("Your IP address is %s" % ip)
    return ip


def user_mac(request):
    mac = get_mac_address(ip = user_ip(request))
    #':'.join(("%012X" % mac)[i:i+2] for i in range(0, 12, 2)
    return mac
    #return HttpResponse("Your mac address is %s" % mac)